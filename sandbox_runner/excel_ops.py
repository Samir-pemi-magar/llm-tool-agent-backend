"""
The actual openpyxl logic.

This file lives only inside the sandbox image.
The main application never manipulates Excel files directly.
"""

import datetime
import openpyxl


MAX_READ_ROWS = 200


class SheetNotFoundError(Exception):
    """Raised when the requested worksheet does not exist."""


def _serialize(value):
    """
    JSON cannot directly represent date/time objects.
    Convert them to ISO strings before returning them.
    """
    if isinstance(
        value,
        (
            datetime.datetime,
            datetime.date,
            datetime.time,
        ),
    ):
        return value.isoformat()

    return value


def _get_sheet(wb, sheet_name: str):
    """
    Resolve an exact worksheet name.

    The caller must explicitly provide a sheet name.
    We intentionally do not fall back to wb.active because that is how
    multi-sheet workbook mistakes happen.
    """
    if not sheet_name:
        raise SheetNotFoundError(
            "sheet_name is required. "
            f"Available sheets: {wb.sheetnames}"
        )

    if sheet_name not in wb.sheetnames:
        raise SheetNotFoundError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    return wb[sheet_name]


def _get_headers(ws) -> list:
    """
    Return the first-row headers for a worksheet.
    """
    return [
        cell.value
        for cell in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]


def _find_column_elsewhere(
    wb,
    column: str,
    current_sheet: str,
) -> list[str]:
    """
    Search other worksheets for a missing column.

    This gives the LLM a useful error instead of forcing it to guess
    another synonym for the same column on the wrong sheet.
    """
    hits = []

    for name in wb.sheetnames:
        if name == current_sheet:
            continue

        ws = wb[name]

        try:
            headers = _get_headers(ws)
        except StopIteration:
            continue

        if column in headers:
            hits.append(name)

    return hits


def _coerce_numeric(value):
    """
    Convert obvious numeric strings into real Excel numeric values.

    Examples:
        '2'       -> 2
        '52000'   -> 52000
        '5376.50' -> 5376.5

    Non-numeric values remain unchanged.
    """
    if not isinstance(value, str):
        return value

    stripped = value.strip()

    if stripped == "":
        return value

    try:
        return int(stripped)
    except ValueError:
        pass

    try:
        return float(stripped)
    except ValueError:
        return value


def _first_available_data_row(
    ws,
    headers: list,
    row_data: dict,
) -> int:
    """
    Find the first row where all user-supplied columns are empty.

    This is intentionally different from ws.append().

    In the Sales Log workbook, rows 3-302 contain formulas in calculated
    columns while the user-input columns are blank. Those rows are valid
    locations for new sales.

    Example:
        Sale Date
        Model Name
        Units Sold
        Sale Price/Unit (USD)

    If all of those supplied fields are empty on a row, that row can be
    used without overwriting an existing record.

    If no available row is found, return max_row + 1.
    """
    supplied_columns = [
        headers.index(header) + 1
        for header in row_data
        if header in headers
    ]

    if not supplied_columns:
        return ws.max_row + 1

    for row_number in range(2, ws.max_row + 1):
        values = [
            ws.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in supplied_columns
        ]

        if all(
            value is None or value == ""
            for value in values
        ):
            return row_number

    return ws.max_row + 1


def list_sheets(file_path: str) -> dict:
    """
    Return every worksheet name and the active worksheet.
    """
    wb = openpyxl.load_workbook(
        file_path,
        read_only=True,
    )

    return {
        "sheets": wb.sheetnames,
        "active_sheet": wb.active.title,
    }


def read_excel_data(
    file_path: str,
    sheet_name: str,
) -> dict:
    """
    Read rows from one specific worksheet.
    """
    wb = openpyxl.load_workbook(
        file_path,
        data_only=True,
    )

    try:
        ws = _get_sheet(
            wb,
            sheet_name,
        )
    except SheetNotFoundError as e:
        return {
            "error": str(e),
        }

    total_rows = ws.max_row
    truncated = total_rows > MAX_READ_ROWS

    rows = [
        [
            _serialize(cell.value)
            for cell in row
        ]
        for row in ws.iter_rows(
            max_row=min(
                total_rows,
                MAX_READ_ROWS,
            )
        )
    ]

    result = {
        "sheet": ws.title,
        "rows": rows,
    }

    if truncated:
        result["truncated"] = True
        result["total_rows"] = total_rows
        result["note"] = (
            f"Only the first {MAX_READ_ROWS} of {total_rows} rows are shown. "
            "Rows beyond this cutoff are NOT visible here and may contain "
            "additional non-empty data. Do not conclude a column is sparse, "
            "incomplete, or mostly empty based on this partial view. For any "
            "total, count, average, min, or max, use summarize_column, which "
            "scans the entire sheet regardless of this cutoff. Use "
            "get_rows_by_value for targeted lookups."
        )

    return result


def update_excel_data(
    file_path: str,
    cell: str,
    value: str,
    sheet_name: str,
) -> dict:
    """
    Update one exact existing cell.
    """
    wb = openpyxl.load_workbook(file_path)

    try:
        ws = _get_sheet(
            wb,
            sheet_name,
        )
    except SheetNotFoundError as e:
        return {
            "error": str(e),
        }

    ws[cell] = _coerce_numeric(value)

    wb.save(file_path)

    return {
        "status": "ok",
        "sheet": ws.title,
        "updated": cell,
        "value": value,
    }


def append_row_data(
    file_path: str,
    row_data: dict,
    sheet_name: str,
) -> dict:
    """
    Add a new record to a specific worksheet.

    row_data keys must exactly match worksheet headers.

    Unlike ws.append(), this function searches for the first row where
    all supplied/input columns are empty. This allows a workbook to have
    pre-filled formulas in calculated columns without forcing new records
    below legends, notes, or unrelated rows later in the worksheet.
    """
    if not isinstance(row_data, dict):
        return {
            "error": "row_data must be an object mapping column headers to values."
        }

    if not row_data:
        return {
            "error": "row_data cannot be empty."
        }

    wb = openpyxl.load_workbook(file_path)

    try:
        ws = _get_sheet(
            wb,
            sheet_name,
        )
    except SheetNotFoundError as e:
        return {
            "error": str(e),
        }

    headers = _get_headers(ws)

    unknown_columns = [
        key
        for key in row_data
        if key not in headers
    ]

    if unknown_columns:
        return {
            "error": (
                f"Unknown column(s): {unknown_columns}. "
                f"Available columns on sheet '{ws.title}': {headers}"
            )
        }

    target_row = _first_available_data_row(
        ws=ws,
        headers=headers,
        row_data=row_data,
    )

    written_values = {}

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        if header not in row_data:
            continue

        value = _coerce_numeric(
            row_data[header]
        )

        ws.cell(
            row=target_row,
            column=column_number,
            value=value,
        )

        written_values[header] = value

    wb.save(file_path)

    return {
        "status": "ok",
        "sheet": ws.title,
        "row_added": target_row,
        "values": written_values,
    }


def summarize_column(
    file_path: str,
    column: str,
    operation: str,
    sheet_name: str,
) -> dict:
    """
    Calculate a summary statistic for one exact column.
    """
    wb = openpyxl.load_workbook(
        file_path,
        data_only=True,
    )

    try:
        ws = _get_sheet(
            wb,
            sheet_name,
        )
    except SheetNotFoundError as e:
        return {
            "error": str(e),
        }

    headers = _get_headers(ws)

    if column not in headers:
        hits = _find_column_elsewhere(
            wb=wb,
            column=column,
            current_sheet=ws.title,
        )

        message = (
            f"Column '{column}' not found on sheet '{ws.title}'. "
            f"Available columns here: {headers}."
        )

        if hits:
            message += (
                f" This column exists on other sheet(s): {hits}. "
                "Retry with sheet_name set to one of those sheets."
            )
        else:
            message += (
                " The column does not appear on another sheet. "
                "Check the exact column name and worksheet headers."
            )

        return {
            "error": message,
        }

    column_index = headers.index(column)

    raw_values = [
        row[column_index]
        for row in ws.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[column_index] is not None
        and row[column_index] != ""
    ]

    if any(
        isinstance(
            value,
            (
                datetime.datetime,
                datetime.date,
                datetime.time,
            ),
        )
        for value in raw_values
    ):
        return {
            "error": (
                f"Column '{column}' contains dates, which cannot be used for "
                f"the '{operation}' numeric operation."
            )
        }

    values = [
        _coerce_numeric(value)
        for value in raw_values
    ]

    non_numeric = [
        value
        for value in values
        if not isinstance(
            value,
            (int, float),
        )
    ]

    if non_numeric:
        return {
            "error": (
                f"Column '{column}' contains non-numeric value(s), for example "
                f"{non_numeric[0]!r}. Numeric summary operations require a "
                "numeric column."
            )
        }

    if operation == "sum":
        result = sum(values)

    elif operation == "count":
        result = len(values)

    elif operation == "average":
        result = (
            sum(values) / len(values)
            if values
            else 0
        )

    elif operation == "min":
        result = (
            min(values)
            if values
            else None
        )

    elif operation == "max":
        result = (
            max(values)
            if values
            else None
        )

    else:
        return {
            "error": (
                f"Unknown operation '{operation}'. "
                "Use sum, count, average, min, or max."
            )
        }

    return {
        "sheet": ws.title,
        "column": column,
        "operation": operation,
        "result": result,
    }


def get_rows_by_value(
    file_path: str,
    column: str,
    values: list[str],
    sheet_name: str,
) -> dict:
    """
    Return rows where a named column contains one of the requested values.
    """
    wb = openpyxl.load_workbook(
        file_path,
        data_only=True,
    )

    try:
        ws = _get_sheet(
            wb,
            sheet_name,
        )
    except SheetNotFoundError as e:
        return {
            "error": str(e),
        }

    headers = _get_headers(ws)

    if column not in headers:
        hits = _find_column_elsewhere(
            wb=wb,
            column=column,
            current_sheet=ws.title,
        )

        message = (
            f"Column '{column}' not found on sheet '{ws.title}'. "
            f"Available columns here: {headers}."
        )

        if hits:
            message += (
                f" This column exists on other sheet(s): {hits}. "
                "Retry with sheet_name set to one of those sheets."
            )
        else:
            message += (
                " It does not appear on another sheet. "
                "Check the exact worksheet headers."
            )

        return {
            "error": message,
        }

    if not isinstance(values, list) or not values:
        return {
            "error": "values must be a non-empty list."
        }

    column_index = headers.index(column)

    matches = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        target_value = row[column_index]

        if target_value is None:
            continue

        target_text = str(target_value).lower()

        if any(
            str(value).lower() in target_text
            for value in values
        ):
            match = {}

            for header, value in zip(
                headers,
                row,
            ):
                if isinstance(header, str) and header.strip():
                    match[header] = _serialize(value)

            matches.append(match)

    return {
        "sheet": ws.title,
        "matches": matches,
    }


REGISTRY = {
    "list_sheets": list_sheets,
    "read_excel_data": read_excel_data,
    "update_excel_data": update_excel_data,
    "append_row_data": append_row_data,
    "summarize_column": summarize_column,
    "get_rows_by_value": get_rows_by_value,
}