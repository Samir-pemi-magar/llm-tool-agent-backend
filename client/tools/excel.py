import openpyxl

def read_excel_data(file_path: str, sheet_name: str | None = None) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        return {"sheet": ws.title, "rows": rows}
    except Exception as e:
        return {"error": str(e)}

def update_excel_data(file_path: str, cell: str, value: str, sheet_name: str | None = None) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[cell] = value
        wb.save(file_path)
        return {"status": "ok", "updated": cell, "value": value}
    except Exception as e:
        return {"error": str(e)}

def summarize_column(file_path: str, column: str, operation: str, sheet_name: str | None = None) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        if column not in headers:
            return {"error": f"Column '{column}' not found. Available columns: {headers}"}
        idx = headers.index(column)

        values = [
            row[idx] for row in ws.iter_rows(min_row=2, values_only=True)
            if row[idx] is not None
        ]

        if operation == "sum":
            result = sum(values)
        elif operation == "count":
            result = len(values)
        elif operation == "average":
            result = sum(values) / len(values) if values else 0
        elif operation == "min":
            result = min(values)
        elif operation == "max":
            result = max(values)
        else:
            return {"error": f"Unknown operation: {operation}. Use sum, count, average, min, or max."}

        return {"column": column, "operation": operation, "result": result}
    except Exception as e:
        return {"error": str(e)}

SCHEMAS = [
    {
        "type": "function",
        "function": {
            "name": "read_excel_data",
            "description": "Read all rows from an Excel workbook sheet",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Path to the .xlsx file"},
                    "sheet_name": {"type": "string", "description": "Optional sheet name; defaults to active sheet"},
                },
                "required": ["file_path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_excel_data",
            "description": "Update a single cell in an Excel workbook and save it",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "cell": {"type": "string", "description": "Cell reference, e.g. 'B2'"},
                    "value": {"type": "string"},
                    "sheet_name": {"type": "string"},
                },
                "required": ["file_path", "cell", "value"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "summarize_column",
            "description": "Compute a summary statistic (sum, count, average, min, or max) over a named column in an Excel sheet. Use this instead of read_excel_data when the user asks for a total, count, or average, since it computes an exact number instead of relying on reading raw rows.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": "Path to the .xlsx file"},
                    "column": {"type": "string", "description": "Exact column header name, e.g. 'Units Sold'"},
                    "operation": {
                        "type": "string",
                        "enum": ["sum", "count", "average", "min", "max"],
                        "description": "The statistic to compute",
                    },
                    "sheet_name": {"type": "string", "description": "Optional sheet name; defaults to active sheet"},
                },
                "required": ["file_path", "column", "operation"],
            },
        },
    },
]

REGISTRY = {
    "read_excel_data": read_excel_data,
    "update_excel_data": update_excel_data,
    "summarize_column": summarize_column,
}